import base64
import io
import json
import os
import tempfile
import threading
import time
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = FastAPI(title="Investigation Board")

DATA_DIR = Path(__file__).parent / "data"
DATA_FILE = DATA_DIR / "board.json"
DATA_DIR.mkdir(exist_ok=True)
SAVE_LOCK = threading.Lock()


class CardModel(BaseModel):
    id: str = ""
    type: str = "text"
    title: str = ""
    body: str = "<p></p>"
    x: float = 200
    y: float = 200
    width: float = 320
    height: float = 220
    tags: list[str] = Field(default_factory=list)
    tagProperties: dict[str, dict[str, Any]] = Field(default_factory=dict)
    sourceUrl: str = ""
    imageUrl: str = ""
    status: str = "active"
    metadata: dict[str, Any] = Field(default_factory=dict)
    tableData: dict[str, Any] = Field(default_factory=lambda: {"columns": [], "rows": []})
    createdAt: str = ""
    updatedAt: str = ""


class EdgeModel(BaseModel):
    id: str = ""
    from_: str = ""
    to: str = ""
    fromPort: str | None = None
    toPort: str | None = None
    label: str = ""
    type: str = ""


class BoardData(BaseModel):
    boardTitle: str = "Research Workspace"
    cards: list[CardModel] = Field(default_factory=list)
    edges: list[EdgeModel] = Field(default_factory=list)
    tagDefinitions: list[dict[str, Any]] = Field(default_factory=list)
    savedViews: list[dict[str, Any]] = Field(default_factory=list)
    panX: float = 0
    panY: float = 0
    zoom: float = 1


class CanvasModel(BaseModel):
    id: str = ""
    name: str = "Canvas"
    cards: list[CardModel] = Field(default_factory=list)
    edges: list[EdgeModel] = Field(default_factory=list)
    tagDefinitions: list[dict[str, Any]] = Field(default_factory=list)
    savedViews: list[dict[str, Any]] = Field(default_factory=list)
    activeQuickView: str = "all"
    panX: float = 0
    panY: float = 0
    zoom: float = 1


class BookData(BaseModel):
    version: int = 2
    bookTitle: str = "Research Workspace"
    activeCanvasId: str = ""
    canvases: list[CanvasModel] = Field(default_factory=list)


def _load_book() -> BookData:
    if DATA_FILE.exists():
        raw: dict[str, Any] = json.loads(DATA_FILE.read_text(encoding="utf-8"))
        if "version" in raw and raw["version"] == 2 and "canvases" in raw:
            for canvas in raw.get("canvases", []):
                for edge in canvas.get("edges", []):
                    if "from" in edge:
                        edge["from_"] = edge.pop("from")
            return BookData(**raw)
        else:
            for edge in raw.get("edges", []):
                if "from" in edge:
                    edge["from_"] = edge.pop("from")
            board = BoardData(**raw)
            cid = "canvas_1"
            canvas = CanvasModel(
                id=cid,
                name=board.boardTitle,
                cards=board.cards,
                edges=board.edges,
                tagDefinitions=board.tagDefinitions,
                savedViews=board.savedViews,
                activeQuickView="all",
                panX=board.panX,
                panY=board.panY,
                zoom=board.zoom,
            )
            return BookData(bookTitle=board.boardTitle, activeCanvasId=cid, canvases=[canvas])
    return BookData()


def _save_book(data: BookData):
    raw: dict[str, Any] = data.model_dump()

    # Before first v2 write, back up any existing v1 file so rollback can
    # restore the original data if needed.
    bak_path = DATA_FILE.with_suffix(".json.v1.bak")
    if not bak_path.exists() and DATA_FILE.exists():
        try:
            existing = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            if "version" not in existing or existing.get("version") != 2:
                DATA_FILE.rename(bak_path)
        except Exception:
            pass

    for canvas in raw.get("canvases", []):
        for edge in canvas.get("edges", []):
            if "from_" in edge:
                edge["from"] = edge.pop("from_")
    for canvas in raw.get("canvases", []):
        for edge in canvas.get("edges", []):
            if "from_" in edge:
                edge["from"] = edge.pop("from_")

    payload = json.dumps(raw, indent=2, ensure_ascii=False)

    # Atomic write + retry to avoid intermittent Windows file-write issues
    # under rapid autosave + reload activity.
    with SAVE_LOCK:
        last_err: OSError | None = None
        for _ in range(3):
            tmp_name = None
            try:
                with tempfile.NamedTemporaryFile(
                    mode="w",
                    encoding="utf-8",
                    delete=False,
                    dir=DATA_DIR,
                    prefix="board_",
                    suffix=".tmp",
                ) as tmp:
                    tmp_name = tmp.name
                    tmp.write(payload)
                    tmp.flush()
                    os.fsync(tmp.fileno())

                os.replace(tmp_name, DATA_FILE)
                return
            except OSError as err:
                last_err = err
                if tmp_name:
                    try:
                        os.unlink(tmp_name)
                    except OSError:
                        pass
                if err.errno in (5, 13, 16, 22):
                    time.sleep(0.05)
                    continue
                raise

        if last_err is not None:
            raise last_err


@app.get("/api/board")
def get_board():
    data = _load_book()
    raw: dict[str, Any] = data.model_dump()
    for canvas in raw.get("canvases", []):
        for edge in canvas.get("edges", []):
            if "from_" in edge:
                edge["from"] = edge.pop("from_")
    return raw


@app.post("/api/board")
def save_board(body: dict[str, Any]):
    raw = body
    for canvas in raw.get("canvases", []):
        for edge in canvas.get("edges", []):
            if "from" in edge:
                edge["from_"] = edge.pop("from")
    data = BookData(**raw)
    _save_book(data)
    return {"ok": True}


@app.get("/api/export/table/{card_id}")
def export_table_xlsx(card_id: str):
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")
    data = _load_book()
    card = None
    for canvas in data.canvases:
        card = next((c for c in canvas.cards if c.id == card_id), None)
        if card:
            break
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    if card.type != "table":
        raise HTTPException(status_code=400, detail="Card is not a table type")

    wb = Workbook()
    ws = wb.active
    ws.title = card.title or "Table"

    td = card.tableData or {"columns": [], "rows": []}
    columns = td.get("columns", [])
    rows = td.get("rows", [])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if columns:
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        from openpyxl.utils import get_column_letter
        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 20
    else:
        ws.cell(row=1, column=1, value="No data")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{card.title or 'table'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/api/download-image")
async def download_image(body: dict[str, Any]):
    url = body.get("url", "")
    if not url or not url.startswith(("http://", "https://")):
        raise HTTPException(status_code=400, detail="Invalid URL")
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=15) as resp:
            content = resp.read()
            content_type = resp.headers.get("Content-Type", "image/png")
            b64 = base64.b64encode(content).decode("utf-8")
            return {"dataUrl": f"data:{content_type};base64,{b64}"}
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


STATIC = Path(__file__).parent.parent / "frontend"
app.mount("/", StaticFiles(directory=str(STATIC), html=True), name="static")